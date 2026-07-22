#!/usr/bin/env python3
"""Importa padron Propietarios Marinas (hojas Agua / Tierra) a cnb_app.

Uso:
  DATABASE_URL=postgresql://cnb_user:cnb_password@127.0.0.1:5433/CNBDB \\
    python3 bin/import_embarcaciones_padron.py \\
    "/home/jorge/Downloads/Propietarios Marinas 01-03-26.xlsx"
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("Instalá openpyxl: pip install openpyxl\n")
    raise

try:
    import psycopg2
    from psycopg2.extras import execute_values
except ImportError:
    sys.stderr.write("Instalá psycopg2-binary: pip install psycopg2-binary\n")
    raise


def parse_num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if text in ("", ".", "-", "-."):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_socio(value):
    n = parse_num(value)
    if n is None:
        return None
    n = int(n)
    return n if 1 <= n <= 99999 else None


def split_nombre(full: str) -> tuple[str, str]:
    full = re.sub(r"\s+", " ", (full or "").strip())
    if not full:
        return ("Sin", "nombre")
    parts = full.split(" ")
    if len(parts) == 1:
        return (parts[0].title(), ".")
    return (parts[0].title(), " ".join(parts[1:]).title())


def clean_text(value, max_len=None):
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    if not text:
        return None
    if max_len:
        return text[:max_len]
    return text


def is_header_row(numero_socio, nombre_barco) -> bool:
    s = str(numero_socio or "").upper()
    n = str(nombre_barco or "").upper()
    return "SOCIO" in s or n in ("EMBARCACION", "NOMBRE", "NOMBRE ")


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 1

    xlsx = Path(sys.argv[1])
    if not xlsx.is_file():
        print(f"No existe: {xlsx}", file=sys.stderr)
        return 1

    dsn = os.environ.get(
        "DATABASE_URL",
        "postgresql://cnb_user:cnb_password@127.0.0.1:5433/CNBDB",
    )
    # quitar query string doctrine
    dsn = dsn.split("?", 1)[0]

    wb = load_workbook(xlsx, read_only=True, data_only=True)

    ubicaciones: dict[tuple[str, str], None] = {("tierra", "Sector tierra"): None, ("agua", "Sin especificar"): None}
    socios: dict[int, tuple[str, str]] = {}
    embarcaciones = []
    seen_matricula: set[str] = set()

    # ---- Agua ----
    ws = wb["Agua "]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        estado = clean_text(row[1], 60)
        numero = parse_socio(row[2])
        propietario = clean_text(row[3], 200)
        tipo = clean_text(row[4], 80) or "velero"
        modelo = clean_text(row[5], 120)
        nombre = clean_text(row[6], 160)
        matricula = clean_text(row[8], 80)
        if is_header_row(row[2], nombre) or not nombre:
            continue
        if tipo.upper() == "DESOCUPADO" and not matricula and numero is None:
            # amarre vacio: guardar con nombre DESOCUPADO si aporta ubicacion
            pass
        ubic_nombre = clean_text(row[17], 120) or "Sin especificar"
        if ubic_nombre.lower() in ("ubicacion", "ubicación"):
            ubic_nombre = "Sin especificar"
        ubicaciones[("agua", ubic_nombre)] = None
        if numero and propietario:
            socios[numero] = split_nombre(propietario)
        if matricula:
            key = matricula.upper()
            if key in seen_matricula:
                matricula = f"{matricula}-{numero or 'x'}"
            seen_matricula.add(matricula.upper())
        embarcaciones.append(
            {
                "ambito": "agua",
                "estado_padron": estado,
                "numero_socio": numero,
                "tipo": tipo,
                "modelo": modelo,
                "nombre": nombre,
                "matricula": matricula,
                "eslora_m": parse_num(row[9]),
                "manga_m": parse_num(row[10]),
                "m2_matricula": parse_num(row[11]),
                "metros_comprados": parse_num(row[13]),
                "paga_expensas_m2": parse_num(row[15]),
                "ubicacion_nombre": ubic_nombre,
                "observaciones": clean_text(row[19]),
                "eslora_medida_m": None,
                "manga_medida_m": None,
                "m2_medidos": None,
            }
        )

    # ---- Tierra ----
    ws = wb["Tierra"]
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row:
            continue
        estado = clean_text(row[1], 60)
        numero = parse_socio(row[2])
        propietario = clean_text(row[3], 200)
        tipo = clean_text(row[4], 80) or "velero"
        modelo = clean_text(row[5], 120)
        nombre = clean_text(row[6], 160)
        matricula = clean_text(row[8], 80)
        if is_header_row(row[2], nombre) or not nombre:
            continue
        ubic_nombre = "Sector tierra"
        ubicaciones[("tierra", ubic_nombre)] = None
        if numero and propietario:
            socios[numero] = split_nombre(propietario)
        if matricula:
            if matricula.upper() in seen_matricula:
                matricula = f"{matricula}-T{numero or 'x'}"
            seen_matricula.add(matricula.upper())
        embarcaciones.append(
            {
                "ambito": "tierra",
                "estado_padron": estado,
                "numero_socio": numero,
                "tipo": tipo,
                "modelo": modelo,
                "nombre": nombre,
                "matricula": matricula,
                "eslora_m": parse_num(row[9]),
                "manga_m": parse_num(row[10]),
                "m2_matricula": parse_num(row[11]),
                "metros_comprados": parse_num(row[17]),
                "paga_expensas_m2": parse_num(row[19]),
                "ubicacion_nombre": ubic_nombre,
                "observaciones": clean_text(row[21]),
                "eslora_medida_m": parse_num(row[13]),
                "manga_medida_m": parse_num(row[14]),
                "m2_medidos": parse_num(row[15]),
            }
        )

    wb.close()

    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    cur = conn.cursor()
    try:
        cur.execute("SET search_path TO cnb_app, public")

        # ubicaciones
        execute_values(
            cur,
            """
            INSERT INTO ubicaciones (ambito, nombre)
            VALUES %s
            ON CONFLICT (ambito, nombre) DO NOTHING
            """,
            [(a, n) for (a, n) in sorted(ubicaciones.keys())],
        )

        cur.execute("SELECT id, ambito, nombre FROM ubicaciones")
        ubic_map = {(ambito, nombre): uid for uid, ambito, nombre in cur.fetchall()}

        # estados_padron
        estados = sorted({e["estado_padron"] for e in embarcaciones if e.get("estado_padron")})
        if estados:
            execute_values(
                cur,
                """
                INSERT INTO estados_padron (nombre)
                VALUES %s
                ON CONFLICT (nombre) DO NOTHING
                """,
                [(n,) for n in estados],
            )
        cur.execute("SELECT id, nombre FROM estados_padron")
        estado_map = {nombre: eid for eid, nombre in cur.fetchall()}

        # socios (upsert por numero_socio)
        for numero, (apellido, nombre) in socios.items():
            cur.execute(
                """
                INSERT INTO socios (numero_socio, nombre, apellido, estado)
                VALUES (%s, %s, %s, 'activo')
                ON CONFLICT (numero_socio) DO UPDATE
                  SET nombre = EXCLUDED.nombre,
                      apellido = EXCLUDED.apellido,
                      updated_at = NOW()
                """,
                (numero, nombre, apellido),
            )

        # limpiar embarcaciones previas del import (rebuild ya las borro, por si re-ejecutan)
        cur.execute("UPDATE tareas SET embarcacion_id = NULL WHERE embarcacion_id IS NOT NULL")
        cur.execute("DELETE FROM reservas_varadero")
        cur.execute("DELETE FROM embarcaciones")

        rows = []
        skipped_fk = 0
        for e in embarcaciones:
            uid = ubic_map.get((e["ambito"], e["ubicacion_nombre"]))
            if uid is None:
                uid = ubic_map.get((e["ambito"], "Sin especificar")) or ubic_map.get((e["ambito"], "Sector tierra"))
            numero = e["numero_socio"]
            # si no existe socio, dejar null (FK)
            if numero is not None and numero not in socios:
                # puede existir en DB de antes
                cur.execute("SELECT 1 FROM socios WHERE numero_socio = %s", (numero,))
                if cur.fetchone() is None:
                    skipped_fk += 1
                    numero = None
            estado_id = estado_map.get(e["estado_padron"]) if e.get("estado_padron") else None
            rows.append(
                (
                    e["ambito"],
                    numero,
                    e["tipo"],
                    e["modelo"],
                    e["nombre"],
                    e["matricula"],
                    e["eslora_m"],
                    e["manga_m"],
                    e["m2_matricula"],
                    e["metros_comprados"],
                    e["paga_expensas_m2"],
                    uid,
                    e["observaciones"],
                    e["eslora_medida_m"],
                    e["manga_medida_m"],
                    e["m2_medidos"],
                    "activa",
                    estado_id,
                    False,
                )
            )

        execute_values(
            cur,
            """
            INSERT INTO embarcaciones (
                ambito, numero_socio, tipo, modelo, nombre, matricula,
                eslora_m, manga_m, m2_matricula, metros_comprados, paga_expensas_m2,
                ubicacion_id, observaciones, eslora_medida_m, manga_medida_m, m2_medidos,
                estado, estado_padron_id, es_cnb
            ) VALUES %s
            """,
            rows,
        )

        conn.commit()
        cur.execute("SELECT count(*) FROM ubicaciones")
        n_ubic = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM embarcaciones")
        n_emb = cur.fetchone()[0]
        cur.execute("SELECT count(*) FROM socios")
        n_soc = cur.fetchone()[0]
        print(f"OK ubicaciones={n_ubic} embarcaciones={n_emb} socios={n_soc} (numero_socio null por FK: {skipped_fk})")
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
